"""
Export utilities for Excel and PDF generation.
Supports styled Excel exports and PDF reports with Chinese font.
"""
import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_dataframe_to_excel(df, sheet_name="Sheet1", title=None):
    """Export a DataFrame to a styled Excel bytes buffer.

    Args:
        df: DataFrame to export.
        sheet_name: Excel sheet name.
        title: Optional title row above the data.

    Returns:
        BytesIO buffer containing the Excel file.
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        start_row = 0

        if title:
            # Write title in row 0
            title_df = pd.DataFrame([[title]])
            title_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=0)
            start_row = 2

        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)

        # Style the workbook
        ws = writer.sheets[sheet_name]
        _style_worksheet(ws, df, start_row, title)

    buffer.seek(0)
    return buffer


def export_multi_sheet_excel(sheets_dict):
    """Export multiple DataFrames as separate sheets in one Excel file.

    Args:
        sheets_dict: dict of {sheet_name: DataFrame}.

    Returns:
        BytesIO buffer.
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _style_worksheet(ws, df, start_row=0)
    buffer.seek(0)
    return buffer


def _style_worksheet(ws, df, start_row=0, title=None):
    """Apply professional styling to a worksheet."""
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="Calibri", size=10)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    even_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    high_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    # Title row
    if title:
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F77B4")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    # Header row
    header_row = start_row + 1
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx in range(start_row + 2, start_row + 2 + len(df)):
        is_even = (row_idx - start_row) % 2 == 0
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Zebra striping
            if is_even:
                cell.fill = even_fill

            # Conditional coloring for known value patterns
            val = str(cell.value) if cell.value else ""
            if val == "Pass":
                cell.fill = pass_fill
            elif val == "Fail":
                cell.fill = fail_fill
                cell.font = Font(name="Calibri", size=10, bold=True, color="D62728")
            elif val == "Conditional":
                cell.fill = warn_fill
            elif val in ("High", "高"):
                cell.fill = high_fill
                cell.font = Font(name="Calibri", size=10, bold=True, color="D62728")

    # Auto-adjust column widths
    for col_idx in range(1, len(df.columns) + 1):
        max_len = len(str(df.columns[col_idx - 1])) + 2
        for row_idx in range(start_row + 2, min(start_row + 52, start_row + 2 + len(df))):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                cell_len = len(str(cell_val))
                max_len = max(max_len, min(cell_len + 2, 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len

    # Freeze header row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def generate_pdf_report(title, content_sections, lang="zh"):
    """Generate a PDF report from structured content sections.

    Args:
        title: Report title string.
        content_sections: List of dicts: [{heading, body}].
        lang: 'zh' or 'en'.

    Returns:
        BytesIO buffer containing the PDF, or None on failure.
    """
    try:
        from fpdf import FPDF
    except ImportError:
        return None

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Use built-in font for English; for Chinese we need a Unicode font
    if lang == "zh":
        # Try to add a Chinese-capable font
        _add_chinese_font(pdf)
    else:
        pdf.set_font("Helvetica", size=12)

    # Title
    pdf.set_font(pdf.font_family, "B", 18)
    pdf.cell(0, 15, title, new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)

    # Content sections
    for section in content_sections:
        heading = section.get("heading", "")
        body = section.get("body", "")

        # Section heading
        pdf.set_font(pdf.font_family, "B", 13)
        pdf.cell(0, 10, heading, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        # Section body
        pdf.set_font(pdf.font_family, "", 10)
        # Handle markdown-like content
        for line in body.split("\n"):
            line = line.strip()
            if not line:
                pdf.ln(3)
                continue
            if line.startswith("- ") or line.startswith("* "):
                pdf.cell(8)
                pdf.multi_cell(0, 6, line)
            else:
                pdf.multi_cell(0, 6, line)

        pdf.ln(5)

    buffer = io.BytesIO()
    buffer.write(pdf.output())
    buffer.seek(0)
    return buffer


def _add_chinese_font(pdf):
    """Try to add a Chinese font to the PDF, fallback to Helvetica."""
    import os

    # Common Chinese font paths on Windows
    font_paths = [
        os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "msyh.ttc"),  # Microsoft YaHei
        os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "simhei.ttf"),  # SimHei
        os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "simsun.ttc"),  # SimSun
    ]

    for fpath in font_paths:
        if os.path.exists(fpath):
            try:
                pdf.add_font("ChineseFont", "", fpath, uni=True)
                pdf.add_font("ChineseFont", "B", fpath, uni=True)
                pdf.set_font("ChineseFont", size=12)
                return
            except Exception:
                continue

    # Fallback: use Helvetica (Chinese chars won't render correctly)
    pdf.set_font("Helvetica", size=12)
